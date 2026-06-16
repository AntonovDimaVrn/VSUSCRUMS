from __future__ import annotations

from dataclasses import dataclass, field
from datetime import UTC, date, datetime
from decimal import Decimal
from io import BytesIO
from pathlib import Path
from typing import Any

from fastapi import UploadFile
from openpyxl import load_workbook
from sqlalchemy import select
from sqlalchemy.orm import Session, selectinload

from app.db.models.project import Project, ProjectMember
from app.db.models.sprint import Sprint
from app.db.models.task import ComplexityClass, QualificationLevel, Task, TaskAssignment, TaskStatus
from app.db.models.upload import Upload, UploadStatus
from app.schemas.upload import ImportTemplateRow

REQUIRED_COLUMNS = [
    "task_id",
    "sprint",
    "story_points",
    "complexity_class",
    "planned_hours",
    "actual_hours",
    "area",
    "participant_name",
    "participant_role",
    "qualification",
    "participant_hours",
]

OPTIONAL_COLUMNS = [
    "task_title",
    "status",
    "external_dependency",
    "sprint_start",
    "sprint_end",
]

TEMPLATE_SAMPLE_ROWS = [
    ImportTemplateRow(
        task_id="SCRUMS-001",
        task_title="Консультация пользователя по работе сервиса",
        sprint="Спринт 1",
        story_points=2,
        complexity_class=ComplexityClass.S,
        planned_hours=18,
        actual_hours=20.1,
        participant_name="Анна Смирнова",
        participant_role="руководитель проекта",
        qualification=QualificationLevel.senior,
        participant_hours=8.4,
        status=TaskStatus.completed,
        area="Консультация",
        external_dependency=False,
        sprint_start=date(2026, 1, 12),
        sprint_end=date(2026, 1, 16),
    ),
    ImportTemplateRow(
        task_id="SCRUMS-001",
        task_title="Консультация пользователя по работе сервиса",
        sprint="Спринт 1",
        story_points=2,
        complexity_class=ComplexityClass.S,
        planned_hours=18,
        actual_hours=20.1,
        participant_name="Иван Иванов",
        participant_role="разработчик",
        qualification=QualificationLevel.middle,
        participant_hours=6.0,
        status=TaskStatus.completed,
        area="Консультация",
        external_dependency=False,
        sprint_start=date(2026, 1, 12),
        sprint_end=date(2026, 1, 16),
    ),
    ImportTemplateRow(
        task_id="SCRUMS-041",
        task_title="Исправление ошибки в заявке пользователя",
        sprint="Спринт 5",
        story_points=3,
        complexity_class=ComplexityClass.M,
        planned_hours=39.5,
        actual_hours=43.1,
        participant_name="Елена Сидорова",
        participant_role="аналитик",
        qualification=QualificationLevel.middle,
        participant_hours=12,
        status=TaskStatus.completed,
        area="Ошибка",
        external_dependency=False,
        sprint_start=date(2026, 2, 9),
        sprint_end=date(2026, 2, 13),
    ),
]

UPLOAD_STORAGE_DIR = Path("/app/storage/uploads")
TEMPLATE_DIR = Path("/app/templates")
TEMPLATE_FILE_PATH = TEMPLATE_DIR / "scrums_input_template.xlsx"


class ExcelImportError(Exception):
    pass


@dataclass
class AssignmentRow:
    participant_name: str
    participant_role: str
    qualification: QualificationLevel
    participant_hours: Decimal


@dataclass
class ParsedTask:
    task_id: str
    sprint_name: str
    story_points: Decimal
    complexity_class: ComplexityClass
    planned_hours: Decimal
    actual_hours: Decimal
    task_title: str | None = None
    status: TaskStatus = TaskStatus.in_progress
    area: str | None = None
    external_dependency: bool = False
    sprint_start: date | None = None
    sprint_end: date | None = None
    assignments: list[AssignmentRow] = field(default_factory=list)


@dataclass
class ImportSummary:
    created_tasks: int
    updated_tasks: int
    assignments_written: int
    sprints_touched: int
    records_count: int


def get_import_template_spec() -> dict[str, Any]:
    return {
        "required_columns": REQUIRED_COLUMNS,
        "optional_columns": OPTIONAL_COLUMNS,
        "sample_rows": [row.model_dump(mode="json") for row in TEMPLATE_SAMPLE_ROWS],
    }


async def import_excel_for_project(
    db: Session,
    project: Project,
    upload_file: UploadFile,
) -> tuple[Upload, ImportSummary]:
    filename = upload_file.filename or "upload.xlsx"
    if not filename.lower().endswith(".xlsx"):
        raise ExcelImportError("Поддерживаются только файлы формата .xlsx.")

    content = await upload_file.read()
    if not content:
        raise ExcelImportError("Файл пустой.")

    upload = Upload(
        project_id=project.id,
        original_filename=filename,
        status=UploadStatus.pending,
        records_count=0,
    )
    db.add(upload)
    db.commit()
    db.refresh(upload)

    storage_path = _persist_upload(project.id, upload.id, filename, content)
    upload.storage_path = storage_path
    db.add(upload)
    db.commit()
    db.refresh(upload)

    try:
        parsed_tasks = _parse_workbook(content)
        summary = _upsert_parsed_tasks(db, project, parsed_tasks)
        upload.status = UploadStatus.processed
        upload.records_count = summary.records_count
        upload.processed_at = datetime.now(UTC)
        upload.error_message = None
        db.add(upload)
        db.commit()
        db.refresh(upload)
        return upload, summary
    except ExcelImportError as exc:
        db.rollback()
        upload.status = UploadStatus.failed
        upload.error_message = str(exc)
        upload.processed_at = datetime.now(UTC)
        db.add(upload)
        db.commit()
        db.refresh(upload)
        raise
    except Exception:
        db.rollback()
        upload.status = UploadStatus.failed
        upload.error_message = "Внутренняя ошибка во время импорта файла."
        upload.processed_at = datetime.now(UTC)
        db.add(upload)
        db.commit()
        db.refresh(upload)
        raise


def _persist_upload(project_id: int, upload_id: int, filename: str, content: bytes) -> str:
    project_dir = UPLOAD_STORAGE_DIR / f"project_{project_id}"
    project_dir.mkdir(parents=True, exist_ok=True)
    safe_name = filename.replace("/", "_")
    output_path = project_dir / f"{upload_id}_{safe_name}"
    output_path.write_bytes(content)
    return str(output_path)


def _parse_workbook(content: bytes) -> list[ParsedTask]:
    workbook = load_workbook(filename=BytesIO(content), data_only=True)
    sheet = workbook.active
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        raise ExcelImportError("В книге нет данных.")

    raw_headers = rows[0]
    headers = [str(cell).strip() if cell is not None else "" for cell in raw_headers]
    if not headers or all(header == "" for header in headers):
        raise ExcelImportError("Не удалось прочитать заголовки колонок.")

    missing_columns = [column for column in REQUIRED_COLUMNS if column not in headers]
    if missing_columns:
        raise ExcelImportError(
            f"В файле отсутствуют обязательные колонки: {', '.join(missing_columns)}."
        )

    header_positions = {header: index for index, header in enumerate(headers)}
    grouped: dict[str, ParsedTask] = {}

    for row_index, row_values in enumerate(rows[1:], start=2):
        if row_values is None or all(value is None or str(value).strip() == "" for value in row_values):
            continue

        row = {header: row_values[index] if index < len(row_values) else None for header, index in header_positions.items()}
        task_id = _get_required_string(row, "task_id", row_index)
        sprint_name = _get_required_string(row, "sprint", row_index)
        story_points = _get_required_decimal(row, "story_points", row_index)
        planned_hours = _get_required_decimal(row, "planned_hours", row_index)
        actual_hours = _get_required_decimal(row, "actual_hours", row_index)
        participant_name = _get_required_string(row, "participant_name", row_index)
        participant_role = _get_required_string(row, "participant_role", row_index)
        participant_hours = _get_required_decimal(row, "participant_hours", row_index)
        complexity_class = _parse_complexity(row.get("complexity_class"), row_index)
        qualification = _parse_qualification(row.get("qualification"), row_index)
        task_status = _parse_status(row.get("status"), row_index)
        external_dependency = _parse_bool(row.get("external_dependency"))
        sprint_start = _parse_date(row.get("sprint_start"), row_index)
        sprint_end = _parse_date(row.get("sprint_end"), row_index)
        task_title = _get_optional_string(row, "task_title")
        area = _get_required_string(row, "area", row_index)

        parsed_row = ParsedTask(
            task_id=task_id,
            sprint_name=sprint_name,
            story_points=story_points,
            complexity_class=complexity_class,
            planned_hours=planned_hours,
            actual_hours=actual_hours,
            task_title=task_title,
            status=task_status,
            area=area,
            external_dependency=external_dependency,
            sprint_start=sprint_start,
            sprint_end=sprint_end,
        )

        existing = grouped.get(task_id)
        if existing is None:
            grouped[task_id] = parsed_row
            existing = grouped[task_id]
        else:
            _assert_task_consistency(existing, parsed_row, row_index)

        existing.assignments.append(
            AssignmentRow(
                participant_name=participant_name,
                participant_role=participant_role,
                qualification=qualification,
                participant_hours=participant_hours,
            )
        )

    parsed_tasks = list(grouped.values())
    if not parsed_tasks:
        raise ExcelImportError("После чтения файла не найдено ни одной строки с заявками.")

    return parsed_tasks


def _upsert_parsed_tasks(db: Session, project: Project, parsed_tasks: list[ParsedTask]) -> ImportSummary:
    sprint_names = sorted({task.sprint_name for task in parsed_tasks})
    existing_sprints = {
        sprint.name: sprint
        for sprint in db.scalars(
            select(Sprint).where(Sprint.project_id == project.id, Sprint.name.in_(sprint_names))
        ).all()
    }

    existing_tasks = {
        task.external_task_id: task
        for task in db.scalars(
            select(Task)
            .where(Task.project_id == project.id, Task.external_task_id.in_([task.task_id for task in parsed_tasks]))
            .options(selectinload(Task.assignments))
        ).all()
    }
    existing_members = {
        member.name: member
        for member in db.scalars(select(ProjectMember).where(ProjectMember.project_id == project.id)).all()
    }

    created_tasks = 0
    updated_tasks = 0
    assignments_written = 0

    sprint_totals: dict[str, Decimal] = {}

    for parsed_task in parsed_tasks:
        sprint = existing_sprints.get(parsed_task.sprint_name)
        if sprint is None:
            sprint = Sprint(
                project_id=project.id,
                name=parsed_task.sprint_name,
                start_date=parsed_task.sprint_start,
                end_date=parsed_task.sprint_end,
            )
            db.add(sprint)
            db.flush()
            existing_sprints[parsed_task.sprint_name] = sprint
        else:
            if parsed_task.sprint_start is not None:
                sprint.start_date = parsed_task.sprint_start
            if parsed_task.sprint_end is not None:
                sprint.end_date = parsed_task.sprint_end

        sprint_totals[parsed_task.sprint_name] = sprint_totals.get(parsed_task.sprint_name, Decimal("0")) + parsed_task.story_points

        task = existing_tasks.get(parsed_task.task_id)
        if task is None:
            task = Task(
                project_id=project.id,
                external_task_id=parsed_task.task_id,
            )
            db.add(task)
            created_tasks += 1
            existing_tasks[parsed_task.task_id] = task
        else:
            task.assignments.clear()
            updated_tasks += 1

        task.sprint = sprint
        task.title = parsed_task.task_title
        task.area = parsed_task.area
        task.story_points = parsed_task.story_points
        task.complexity_class = parsed_task.complexity_class
        task.planned_hours = parsed_task.planned_hours
        task.actual_hours = parsed_task.actual_hours
        task.status = parsed_task.status
        task.external_dependency = parsed_task.external_dependency

        for assignment in parsed_task.assignments:
            member = existing_members.get(assignment.participant_name)
            if member is None:
                member = ProjectMember(
                    project_id=project.id,
                    name=assignment.participant_name,
                    role=assignment.participant_role,
                    qualification=assignment.qualification,
                    capacity_hours=40,
                )
                db.add(member)
                existing_members[assignment.participant_name] = member
            else:
                member.role = assignment.participant_role
                member.qualification = assignment.qualification

            task.assignments.append(
                TaskAssignment(
                    participant_name=assignment.participant_name,
                    qualification=assignment.qualification,
                    participant_hours=assignment.participant_hours,
                )
            )
            assignments_written += 1

    for sprint_name, total_story_points in sprint_totals.items():
        sprint = existing_sprints[sprint_name]
        sprint.planned_story_points = total_story_points

    db.commit()

    return ImportSummary(
        created_tasks=created_tasks,
        updated_tasks=updated_tasks,
        assignments_written=assignments_written,
        sprints_touched=len(sprint_totals),
        records_count=sum(len(task.assignments) for task in parsed_tasks),
    )


def _assert_task_consistency(existing: ParsedTask, incoming: ParsedTask, row_index: int) -> None:
    fields_to_compare = [
        ("sprint", existing.sprint_name, incoming.sprint_name),
        ("story_points", existing.story_points, incoming.story_points),
        ("complexity_class", existing.complexity_class, incoming.complexity_class),
        ("planned_hours", existing.planned_hours, incoming.planned_hours),
        ("actual_hours", existing.actual_hours, incoming.actual_hours),
        ("status", existing.status, incoming.status),
        ("external_dependency", existing.external_dependency, incoming.external_dependency),
    ]
    for field_name, left, right in fields_to_compare:
        if left != right:
            raise ExcelImportError(
                f"Строка {row_index}: у заявки {existing.task_id} поле {field_name} отличается между строками."
            )


def _get_required_string(row: dict[str, Any], key: str, row_index: int) -> str:
    value = row.get(key)
    if value is None or str(value).strip() == "":
        raise ExcelImportError(f"Строка {row_index}: колонка {key} обязательна.")
    return str(value).strip()


def _get_optional_string(row: dict[str, Any], key: str) -> str | None:
    value = row.get(key)
    if value is None:
        return None
    stringified = str(value).strip()
    return stringified or None


def _get_required_decimal(row: dict[str, Any], key: str, row_index: int) -> Decimal:
    value = row.get(key)
    if value is None or str(value).strip() == "":
        raise ExcelImportError(f"Строка {row_index}: колонка {key} обязательна.")
    try:
        decimal_value = Decimal(str(value))
    except Exception as exc:  # pragma: no cover - protective validation
        raise ExcelImportError(f"Строка {row_index}: колонка {key} должна быть числом.") from exc
    if decimal_value <= 0:
        raise ExcelImportError(f"Строка {row_index}: колонка {key} должна быть больше 0.")
    return decimal_value


def _parse_complexity(value: Any, row_index: int) -> ComplexityClass:
    try:
        return ComplexityClass(str(value).strip())
    except Exception as exc:
        raise ExcelImportError(
            f"Строка {row_index}: complexity_class должен быть одним из S, M, L, XL."
        ) from exc


def _parse_qualification(value: Any, row_index: int) -> QualificationLevel:
    try:
        return QualificationLevel(str(value).strip().lower())
    except Exception as exc:
        raise ExcelImportError(
            f"Строка {row_index}: qualification должен быть одним из junior, middle, senior."
        ) from exc


def _parse_status(value: Any, row_index: int) -> TaskStatus:
    if value is None or str(value).strip() == "":
        return TaskStatus.in_progress
    try:
        return TaskStatus(str(value).strip().lower())
    except Exception as exc:
        raise ExcelImportError(
            f"Строка {row_index}: status должен быть одним из completed, in_progress, blocked."
        ) from exc


def _parse_bool(value: Any) -> bool:
    if value is None or str(value).strip() == "":
        return False
    if isinstance(value, bool):
        return value
    normalized = str(value).strip().lower()
    return normalized in {"1", "true", "yes", "y", "да"}


def _parse_date(value: Any, row_index: int) -> date | None:
    if value is None or str(value).strip() == "":
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    try:
        return date.fromisoformat(str(value).strip())
    except ValueError as exc:
        raise ExcelImportError(
            f"Строка {row_index}: дата должна быть в формате YYYY-MM-DD."
        ) from exc
